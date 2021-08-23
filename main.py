# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import re
import difflib
import time
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from concurrent.futures import ThreadPoolExecutor

attach_window = None

thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")

wb_map = openpyxl.load_workbook('data/map.xlsx')
ws_map_owner = wb_map['单位']
ws_map_name = wb_map['名称']
ws_map_unit = wb_map['规格']
ws_map_factory = wb_map['厂家']

if wb_map.__contains__('相似名称'):
    ws_same_name = wb_map['相似名称']
    wb_map.remove(ws_same_name)
ws_same_name = wb_map.create_sheet('相似名称')
ws_same_name.append(['要求名称', '数据名称'])
for cell in ws_same_name[1]:
    cell.fill = PatternFill("solid", fgColor="47A1A1")

if wb_map.__contains__('相似单位'):
    ws_same_owner = wb_map['相似单位']
    wb_map.remove(ws_same_owner)
ws_same_owner = wb_map.create_sheet('相似单位')
ws_same_owner.append(['要求单位', '数据单位'])
for cell in ws_same_owner[1]:
    cell.fill = PatternFill("solid", fgColor="47A1A1")


wb = openpyxl.load_workbook('data/data.xlsx')
ws_p = wb['要求']
ws_data = wb['原始数据']

if wb.__contains__('筛选数据'):
    ws_filter = wb['筛选数据']
    wb.remove(ws_filter)
ws_filter = wb.create_sheet('筛选数据')
title_list = ['单位2', '客户分类', '负责人', '单价', '金额', '报价', '销售额（报价）', '限价', '销售额（限价）', '代理商']
data_title_row = ws_data[1]
for cell in reversed(data_title_row):
    title_list.insert(0, cell.value)
ws_filter.append(tuple(title_list))
for cell in ws_filter[1]:
    cell.fill = PatternFill("solid", fgColor="47A1A1")

if wb.__contains__('未筛选数据'):
    ws_un_filter = wb['未筛选数据']
    wb.remove(ws_un_filter)
ws_un_filter = wb.create_sheet('未筛选数据')
ws_un_filter.append(tuple(title_list))
for cell in ws_un_filter[1]:
    cell.fill = PatternFill("solid", fgColor="47A1A1")

total = ws_data.max_row - 1
count = 0


def load_excel():
    global count
    print_log('开始处理...')
    count = 0
    # with ThreadPoolExecutor(100) as t:
    #     for data_row in ws_data.iter_rows(min_row=2, values_only=False):
    #         t.submit(filter_data, data_row)

    for data_row in ws_data.iter_rows(min_row=2, values_only=False):
        filter_data(data_row)

    print_log('数据处理完成...\n共%d条记录，筛选%d条记录' % (total,count))
    wb.save('data/data.xlsx')
    wb_map.save('data/map.xlsx')


def filter_data(data_row):
    global count
    if attach_window is not None:
        attach_window.update_progress(count, total)
    data_value_list = list(map(lambda v: v.value, data_row))
    data_name = data_row[2].value
    data_unit = data_row[3].value
    data_factory = data_row[5].value
    data_owner = data_row[6].value
    data_num = data_row[7].value or 0
    selected = False

    for row in ws_p.iter_rows(min_row=2, values_only=True):
        product_name = row[0]
        product_unit = row[1]
        product_factory = row[2]
        product_owner = row[4]
        product_ext = row[5:11]
        # print_log('product name: %s, product owner: %s' % (product_name, product_owner))
        # print_log('data name: %s, data owner: %s' % (data_name, data_owner))
        ratio = difflib.SequenceMatcher(None, product_name, data_name).quick_ratio()
        product_owner_suffix = re.sub(r'.*市|.*县|医院|卫生院', '', product_owner)
        data_owner_suffix = re.sub(r'.*市|.*县|医院|卫生院', '', data_owner)
        # print_log('product_owner_suffix: %s, data_owner_suffix: %s' % (product_owner_suffix, data_owner_suffix))
        ratio2 = difflib.SequenceMatcher(None, product_owner_suffix, data_owner_suffix).quick_ratio()
        # print_log('%s  %s' % (ratio, ratio2))

        is_owner_equal = is_equal(ws_map_owner, product_owner, data_owner)
        is_name_equal = is_equal(ws_map_name, product_name, data_name)
        is_unit_equal = is_equal(ws_map_unit, product_unit, data_unit)
        is_factory_equal = is_equal(ws_map_factory, product_factory, data_factory)

        if ratio2 > 0.6 and not is_owner_equal:
            print_log('异常数据：%s' % data_value_list)
            data_row[6].fill = PatternFill("solid", fgColor="FF0000")
            ws_same_name.append([product_owner, data_owner])

        if is_name_equal and is_owner_equal and is_unit_equal and is_factory_equal:
            amount1 = (product_ext[2] or 0) * data_num
            amount2 = (product_ext[3] or 0) * data_num
            amount3 = (product_ext[4] or 0) * data_num
            product_ext_list = list(product_ext)
            product_ext_list.insert(5, amount3)
            product_ext_list.insert(4, amount2)
            product_ext_list.insert(3, amount1)
            product_ext_list.insert(0, row[3])

            data_value_list.extend(product_ext_list)
            ws_filter.append(tuple(data_value_list))
            print_log(data_value_list)
            selected = True
            count += 1
            break

    if not selected:
        ws_un_filter.append(data_value_list)


def uni_product_name():
    print_log('开始统一筛选列表数据...')
    for row in ws_filter.iter_rows(min_row=2):
        owner = row[6].value
        for r in ws_map_owner.iter_rows(min_row=2, values_only=True):
            if list(r).__contains__(owner):
                default_name = r[0]
                if default_name != owner:
                    print_log('统一厂家：%s -> %s' % (owner, default_name))
                    row[6].value = default_name

        name = row[2].value
        for r in ws_map_name.iter_rows(min_row=2, values_only=True):
            if list(r).__contains__(name):
                default_name = r[0]
                if default_name != name:
                    print_log('统一单位：%s -> %s' % (name, default_name))
                    row[2].value = default_name

        unit = row[3].value
        for r in ws_map_unit.iter_rows(min_row=2, values_only=True):
            if list(r).__contains__(unit):
                default_name = r[0]
                if default_name != unit:
                    print_log('统一规格：%s -> %s' % (unit, default_name))
                    row[3].value = default_name

        factory = row[5].value
        for r in ws_map_factory.iter_rows(min_row=2, values_only=True):
            if list(r).__contains__(factory):
                default_name = r[0]
                if default_name != factory:
                    print_log('统一厂家：%s -> %s' % (factory, default_name))
                    row[5].value = default_name

    print_log('统一筛选列表数据完成...')
    wb.save('data/data.xlsx')


def is_equal(sheet, str1, str2):
    str1.strip()
    str2.strip()

    if str1 == str2:
        return True

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if list(row).__contains__(str1) and list(row).__contains__(str2):
            return True

    return False


def print_log(text):
    print(text)
    if attach_window is not None:
        attach_window.console_log(text)


def start_work(attach=None):
    global attach_window
    attach_window = attach

    time_start = time.time()
    load_excel()
    uni_product_name()
    time_end = time.time()
    duration = time_end - time_start
    print_log('总共用时：%s' % duration)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    start_work()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
